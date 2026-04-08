import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from utils import extrair_patrimonios


# ==================================
# FUNÇÃO: Ler patrimônios do Excell
# ==================================
def carregar_patrimonios_excel(excel_file, sheet_name, coluna):

    # Lê o Excel
    df = pd.read_excel(excel_file, sheet_name=sheet_name, header=1)

    # Verifica se a coluna existe
    if coluna not in df.columns:
        raise Exception(f"Coluna '{coluna}' não encontrada no Excel.")

    # Set evita duplicados
    set_patrimonios = set()

    # Remove células vazias
    valores = df[coluna].dropna().astype(str).tolist()

    # Percorre cada linha
    for val in valores:

        # Extrai patrimônios do texto
        encontrados = extrair_patrimonios(val)

        # Adiciona no set
        for p in encontrados:
            set_patrimonios.add(p)

    return set_patrimonios


# ==================================
# FUNÇÃO: Marcar Excel
# ==================================
def marcar_excel(excel_file, sheet_name, coluna, set_pdf, excel_saida):

    # Abre o Excel
    wb = openpyxl.load_workbook(excel_file)

    # Seleciona aba
    ws = wb[sheet_name]

    # Cor amarela
    amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    linha_cabecalho = 2
    idx_col = None

    # Descobre qual coluna é a de patrimônio
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=linha_cabecalho, column=col).value
        if val and str(val).strip() == coluna:
            idx_col = col
            break

    if not idx_col:
        raise Exception("Coluna de patrimônio não encontrada.")

    max_col = ws.max_column
    divergencias = 0

    # Percorre as linhas da planilha
    for row in range(linha_cabecalho + 1, ws.max_row + 1):

        valor = ws.cell(row=row, column=idx_col).value
        texto = str(valor) if valor else ""

        patrimonios = extrair_patrimonios(texto)

        encontrado = any(p in set_pdf for p in patrimonios)

        # Se não encontrou no PDF
        if patrimonios and not encontrado:

            divergencias += 1

            # Marca a linha inteira
            for c in range(1, max_col + 1):
                ws.cell(row=row, column=c).fill = amarelo

    wb.save(excel_saida)

    print(f"{divergencias} divergências encontradas no Excel.")